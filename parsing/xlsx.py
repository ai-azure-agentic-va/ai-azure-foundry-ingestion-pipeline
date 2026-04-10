"""Excel parser using openpyxl."""

import io
import logging

from openpyxl import load_workbook

from parsing.base import BaseParser, ParseResult

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Extract sheet data from Excel files using openpyxl."""

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".xlsm"]

    def parse(self, file_bytes: bytes) -> ParseResult:
        logger.info(f"[XlsxParser] Parsing XLSX ({len(file_bytes)} bytes)")
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

        sheets_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in row_text):
                    rows.append(" | ".join(row_text))
            if rows:
                sheets_text.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))

        wb.close()
        full_text = "\n\n".join(sheets_text)

        logger.info(f"[XlsxParser] Extracted {len(sheets_text)} sheets, {len(full_text)} chars")
        return ParseResult(
            full_text=full_text,
            pages=[],
            page_count=len(sheets_text),
            metadata={"format": "xlsx", "sheet_count": len(sheets_text)},
        )
